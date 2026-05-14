"""
客户管理服务

封装客户相关的所有业务逻辑：
- 创建客户
- 查询客户列表
- 查询客户详情
- 软删除客户
- 生成客户摘要
"""
import logging
import uuid
from io import BytesIO
from datetime import date, datetime
from typing import Optional, List

from sqlalchemy import case, select, and_, or_, func, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.customer import Customer
from app.models.record import Record
from app.models.user import User
from app.schemas.customer import CustomerCreate, CustomerListItem, CustomerDetail, CustomerListResponse, SummaryGenerateResponse, CustomerChatResponse, AdviceGenerateResponse, CustomerUpdate
from app.ai.kimi_client import KimiClient
from app.core.prompts import (
    customer_summary as _customer_summary_prompt,
    customer_summary_system as _customer_summary_system,
    customer_chat as _customer_chat_prompt,
    customer_chat_system as _customer_chat_system,
    advice as _advice_prompt,
    advice_system as _advice_system,
)
from app.services.customer_advice_store import CustomerAdviceStore

logger = logging.getLogger(__name__)


class CustomerService:
    """
    客户服务类
    
    所有客户相关业务逻辑封装在此类中
    """
    
    def __init__(self, session: AsyncSession):
        """
        初始化服务
        
        Args:
            session: 数据库会话
        """
        self.session = session
        self.advice_store = CustomerAdviceStore()

    async def _get_user_industry_key(self, user_id: str) -> str:
        user = await self.session.get(User, user_id)
        return getattr(user, "industry_key", None) or "generic"
    
    async def create_customer(
        self,
        user_id: str,
        data: CustomerCreate,
    ) -> str:
        """
        创建新客户
        
        Args:
            user_id: 当前用户ID
            data: 创建客户请求数据
            
        Returns:
            新创建客户的ID
        """
        # 生成 UUID
        customer_id = str(uuid.uuid4())

        # 识别地址结构化
        location_normalized = {"location_raw": None, "location_city": None, "location_district": None, "location_subarea": None}
        if data.location:
            from app.services.location_normalizer import LocationNormalizer
            location_normalized = await LocationNormalizer().normalize(data.location)

        # 创建客户实体
        customer = Customer(
            id=customer_id,
            user_id=user_id,
            name=data.name.strip(),
            phone=data.phone.strip() if data.phone else None,
            gender=data.gender.strip() if data.gender else None,
            age=data.age,
            birthday=data.birthday,
            location_raw=location_normalized.get("location_raw"),
            location_city=location_normalized.get("location_city"),
            location_district=location_normalized.get("location_district"),
            location_subarea=location_normalized.get("location_subarea"),
            tags=data.tags if data.tags else [],
            summary_status="stale",
            deleted_at=None,
        )
        
        self.session.add(customer)
        await self.session.flush()
        
        return customer_id

    async def import_customers_excel(self, user_id: str, content: bytes) -> dict:
        """从 Excel 批量导入客户。"""
        from openpyxl import load_workbook

        try:
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("Excel 文件无法读取，请上传 .xlsx 格式文件") from exc

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {
                "created": 0,
                "skipped": 0,
                "failed": 0,
                "total_rows": 0,
                "errors": [{"row": 1, "name": None, "reason": "表格为空"}],
            }

        header_map = self._build_import_header_map(rows[0])
        if "name" not in header_map:
            return {
                "created": 0,
                "skipped": 0,
                "failed": 0,
                "total_rows": max(len(rows) - 1, 0),
                "errors": [{"row": 1, "name": None, "reason": "缺少姓名列"}],
            }

        existing_result = await self.session.execute(
            select(Customer).where(
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None),
            )
        )
        existing_customers = existing_result.scalars().all()
        existing_keys = {
            self._customer_duplicate_key(customer.name, customer.phone)
            for customer in existing_customers
        }
        batch_keys: set[tuple[str, str]] = set()

        created = 0
        skipped = 0
        failed = 0
        total_rows = 0
        errors: list[dict] = []

        for row_index, row in enumerate(rows[1:], start=2):
            if self._is_empty_import_row(row):
                continue
            total_rows += 1
            try:
                item = self._parse_customer_import_row(row, header_map)
                name = item["name"]
                duplicate_key = self._customer_duplicate_key(name, item.get("phone"))
                if duplicate_key in existing_keys or duplicate_key in batch_keys:
                    skipped += 1
                    self._append_import_error(
                        errors,
                        row=row_index,
                        name=name,
                        reason="客户已存在，已跳过",
                    )
                    continue

                data = CustomerCreate(
                    name=name,
                    phone=item.get("phone"),
                    gender=item.get("gender"),
                    age=item.get("age"),
                    birthday=item.get("birthday"),
                    location=item.get("location"),
                    tags=item.get("tags") or [],
                )
                await self.create_customer(user_id=user_id, data=data)
                batch_keys.add(duplicate_key)
                created += 1
            except ValueError as exc:
                failed += 1
                self._append_import_error(
                    errors,
                    row=row_index,
                    name=self._cell_to_text(self._get_import_cell(row, header_map, "name")),
                    reason=str(exc),
                )
            except Exception:
                failed += 1
                logger.exception("Failed to import customer row %s", row_index)
                self._append_import_error(
                    errors,
                    row=row_index,
                    name=self._cell_to_text(self._get_import_cell(row, header_map, "name")),
                    reason="导入失败",
                )

        return {
            "created": created,
            "skipped": skipped,
            "failed": failed,
            "total_rows": total_rows,
            "errors": errors,
        }

    def _append_import_error(
        self,
        errors: list[dict],
        *,
        row: int,
        name: str | None,
        reason: str,
    ) -> None:
        if len(errors) >= 50:
            return
        errors.append({"row": row, "name": name, "reason": reason})

    def _build_import_header_map(self, header_row: tuple) -> dict[str, int]:
        aliases = {
            "name": {"客户姓名", "姓名", "客户", "名字", "name", "customer name"},
            "phone": {"电话", "手机", "手机号", "联系方式", "联系电话", "phone", "mobile"},
            "gender": {"性别", "gender"},
            "age": {"年龄", "age"},
            "birthday": {"生日", "出生日期", "客户生日", "birthday", "birth date"},
            "location": {"地址", "客户地址", "地区", "常住地址", "location", "address"},
            "tags": {"标签", "客户标签", "分类", "备注标签", "tags"},
        }
        normalized_aliases = {
            key: {self._normalize_import_header(alias) for alias in values}
            for key, values in aliases.items()
        }

        header_map: dict[str, int] = {}
        for index, cell in enumerate(header_row or ()):
            header = self._normalize_import_header(cell)
            if not header:
                continue
            for key, values in normalized_aliases.items():
                if key not in header_map and header in values:
                    header_map[key] = index
        return header_map

    def _normalize_import_header(self, value: object) -> str:
        return str(value or "").strip().lower().replace(" ", "").replace("_", "")

    def _is_empty_import_row(self, row: tuple) -> bool:
        return not any(self._cell_to_text(value) for value in row or ())

    def _get_import_cell(self, row: tuple, header_map: dict[str, int], key: str) -> object:
        index = header_map.get(key)
        if index is None or index >= len(row):
            return None
        return row[index]

    def _parse_customer_import_row(self, row: tuple, header_map: dict[str, int]) -> dict:
        name = self._cell_to_text(self._get_import_cell(row, header_map, "name"))
        if not name:
            raise ValueError("姓名不能为空")
        if len(name) > 100:
            raise ValueError("姓名不能超过 100 个字符")

        phone = self._cell_to_text(self._get_import_cell(row, header_map, "phone"))
        if phone and len(phone) > 50:
            raise ValueError("电话不能超过 50 个字符")

        gender = self._cell_to_text(self._get_import_cell(row, header_map, "gender"))
        if gender and len(gender) > 20:
            raise ValueError("性别不能超过 20 个字符")

        location = self._cell_to_text(self._get_import_cell(row, header_map, "location"))
        if location and len(location) > 255:
            raise ValueError("地址不能超过 255 个字符")

        age = self._parse_import_age(self._get_import_cell(row, header_map, "age"))
        birthday = self._parse_import_birthday(
            self._get_import_cell(row, header_map, "birthday")
        )
        tags = self._parse_import_tags(self._get_import_cell(row, header_map, "tags"))
        return {
            "name": name,
            "phone": phone,
            "gender": gender,
            "age": age,
            "birthday": birthday,
            "location": location,
            "tags": tags,
        }

    def _cell_to_text(self, value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            text = str(int(value))
        else:
            text = str(value)
        text = text.strip()
        return text or None

    def _parse_import_age(self, value: object) -> int | None:
        text = self._cell_to_text(value)
        if not text:
            return None
        try:
            age = int(float(text.replace("岁", "").strip()))
        except ValueError as exc:
            raise ValueError("年龄必须是数字") from exc
        if age < 0 or age > 120:
            raise ValueError("年龄必须在 0-120 之间")
        return age

    def _parse_import_birthday(self, value: object) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        text = self._cell_to_text(value)
        if not text:
            return None

        normalized = (
            text.replace("年", "-")
            .replace("月", "-")
            .replace("日", "")
            .replace("/", "-")
            .replace(".", "-")
        )
        try:
            return datetime.strptime(normalized, "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValueError("生日格式应为 YYYY-MM-DD") from exc

    def _parse_import_tags(self, value: object) -> list[str]:
        text = self._cell_to_text(value)
        if not text:
            return []
        import re

        tags = []
        for part in re.split(r"[，,、;；\n]+", text):
            tag = part.strip()
            if tag and tag not in tags:
                tags.append(tag[:40])
        return tags[:20]

    def _customer_duplicate_key(self, name: str, phone: str | None) -> tuple[str, str]:
        return (str(name or "").strip().lower(), str(phone or "").strip())
    
    async def get_customer_list(
        self,
        user_id: str,
        keyword: Optional[str] = None,
        sort_by: Optional[str] = "updated_at",
        sort_order: Optional[str] = "desc",
        page: int = 1,
        page_size: int = 20,
        summary_status: Optional[str] = None,
        stale_contact: bool = False,
    ) -> CustomerListResponse:
        """
        获取客户列表
        """
        from sqlalchemy import asc, desc as desc_order, or_
        from app.services.ai_service import STALE_CONTACT_MONTHS, subtract_months

        # 构建基础查询：未删除 + 属于当前用户
        query = select(Customer).where(
            and_(
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None)
            )
        )

        # 画像状态过滤
        if summary_status:
            statuses = [s.strip() for s in summary_status.split(",") if s.strip()]
            if statuses:
                query = query.where(Customer.summary_status.in_(statuses))

        # 超期未联系过滤
        if stale_contact:
            threshold = subtract_months(datetime.now().date(), STALE_CONTACT_MONTHS)
            threshold_dt = datetime(threshold.year, threshold.month, threshold.day)
            latest_sub = (
                select(Record.customer_id, func.max(Record.created_at).label("last_contact"))
                .group_by(Record.customer_id)
                .subquery()
            )
            query = query.outerjoin(
                latest_sub, latest_sub.c.customer_id == Customer.id
            ).where(
                or_(
                    latest_sub.c.last_contact.is_(None),
                    latest_sub.c.last_contact < threshold_dt,
                )
            )

        # 如果有搜索关键词，添加姓名或标签模糊匹配
        if keyword and keyword.strip():
            keyword_clean = keyword.strip()
            # 使用参数化查询避免 SQL 注入（通过 ilike 操作符）
            search_pattern = f"%{keyword_clean}%"
            query = query.where(
                or_(
                    Customer.name.ilike(search_pattern),
                    # JSON array contains 元素匹配（PostgreSQL jsonb）
                    Customer.tags.contains([keyword_clean])
                )
            )
        
        # 排序
        order_func = desc_order if sort_order == "desc" else asc
        if sort_by == "name":
            query = query.order_by(order_func(Customer.name))
        elif sort_by == "created_at":
            query = query.order_by(order_func(Customer.created_at))
        else:
            # 默认按更新时间
            query = query.order_by(order_func(Customer.updated_at))
        
        count_query = select(func.count()).select_from(query.subquery())
        count_result = await self.session.execute(count_query)
        total = count_result.scalar() or 0

        offset = max(page - 1, 0) * page_size
        paged_query = query.offset(offset).limit(page_size)

        result = await self.session.execute(paged_query)
        customers = result.scalars().all()
        
        # 转换为响应 Schema
        items = [
            CustomerListItem(
                id=c.id,
                name=c.name,
                avatar=c.avatar,
                phone=c.phone,
                gender=c.gender,
                age=c.age,
                birthday=c.birthday,
                location_raw=c.location_raw,
                location_city=c.location_city,
                location_district=c.location_district,
                location_subarea=c.location_subarea,
                tags=c.tags,
                summary_status=c.summary_status,
                updated_at=c.updated_at,
            )
            for c in customers
        ]
        
        return CustomerListResponse(
            items=items,
            total=total,
            page=page,
            page_size=page_size,
        )
    
    async def get_customer_detail(
        self,
        user_id: str,
        customer_id: str,
    ) -> Optional[CustomerDetail]:
        """
        获取客户详情
        
        Args:
            user_id: 当前用户ID
            customer_id: 客户ID
            
        Returns:
            客户详情，如果不存在或已删除则返回 None
        """
        query = select(Customer).where(
            and_(
                Customer.id == customer_id,
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None)
            )
        )
        
        result = await self.session.execute(query)
        customer = result.scalar_one_or_none()
        
        if not customer:
            return None
        
        return CustomerDetail(
            id=customer.id,
            name=customer.name,
            avatar=customer.avatar,
            phone=customer.phone,
            gender=customer.gender,
            age=customer.age,
            birthday=customer.birthday,
            location_raw=customer.location_raw,
            location_city=customer.location_city,
            location_district=customer.location_district,
            location_subarea=customer.location_subarea,
            tags=customer.tags,
            summary_text=customer.summary_text,
            summary_status=customer.summary_status,
            created_at=customer.created_at,
            updated_at=customer.updated_at,
        )
    
    async def delete_customer(
        self,
        user_id: str,
        customer_id: str,
    ) -> bool:
        """
        软删除客户
        
        Args:
            user_id: 当前用户ID
            customer_id: 要删除的客户ID
            
        Returns:
            是否删除成功（客户不存在或已删除返回 False）
        """
        # 先查询确认客户存在且未删除
        query = select(Customer).where(
            and_(
                Customer.id == customer_id,
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None)
            )
        )
        
        result = await self.session.execute(query)
        customer = result.scalar_one_or_none()
        
        if not customer:
            return False
        
        # 软删除：更新 deleted_at 字段
        customer.deleted_at = datetime.utcnow()
        await self.session.flush()
        
        return True

    async def update_customer(
        self,
        user_id: str,
        customer_id: str,
        data: CustomerUpdate,
    ) -> Optional[CustomerDetail]:
        """更新客户资料"""
        query = select(Customer).where(
            and_(
                Customer.id == customer_id,
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None)
            )
        )

        result = await self.session.execute(query)
        customer = result.scalar_one_or_none()

        if not customer:
            return None

        if data.name is not None:
            customer.name = data.name.strip()
        if data.phone is not None:
            customer.phone = data.phone.strip() or None
        if data.gender is not None:
            customer.gender = data.gender.strip() or None
        if data.age is not None:
            customer.age = data.age
        if data.birthday is not None:
            customer.birthday = data.birthday
        if data.tags is not None:
            customer.tags = data.tags
        if data.location is not None:
            from app.services.location_normalizer import LocationNormalizer
            normalized = await LocationNormalizer().normalize(data.location)
            customer.location_raw = normalized.get("location_raw")
            customer.location_city = normalized.get("location_city")
            customer.location_district = normalized.get("location_district")
            customer.location_subarea = normalized.get("location_subarea")

        customer.updated_at = datetime.utcnow()
        await self.session.flush()

        return CustomerDetail(
            id=customer.id,
            name=customer.name,
            avatar=customer.avatar,
            phone=customer.phone,
            gender=customer.gender,
            age=customer.age,
            birthday=customer.birthday,
            location_raw=customer.location_raw,
            location_city=customer.location_city,
            location_district=customer.location_district,
            location_subarea=customer.location_subarea,
            tags=customer.tags,
            summary_text=customer.summary_text,
            summary_status=customer.summary_status,
            created_at=customer.created_at,
            updated_at=customer.updated_at,
        )
    
    async def generate_summary(
        self,
        user_id: str,
        customer_id: str,
    ) -> SummaryGenerateResponse:
        """
        生成客户摘要
        
        基于客户的所有 records，调用 LLM 生成摘要。
        状态流转：stale → updating → ready/failed
        
        Args:
            user_id: 当前用户ID
            customer_id: 客户ID
            
        Returns:
            生成结果
            
        Raises:
            HTTPException: 各种校验失败时抛出
        """
        from fastapi import HTTPException
        
        # 1. 校验客户存在且属于当前用户
        query = select(Customer).where(
            and_(
                Customer.id == customer_id,
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None)
            )
        )
        result = await self.session.execute(query)
        customer = result.scalar_one_or_none()
        
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在或无权访问")
        
        # 2. 查询该客户的所有 records（按时间正序，从旧到新）
        records_query = select(Record).where(
            Record.customer_id == customer_id
        ).order_by(
            Record.created_at  # 正序排列
        )
        
        result = await self.session.execute(records_query)
        records = list(result.scalars().all())
        
        if not records:
            raise HTTPException(status_code=400, detail="该客户暂无沟通记录，无法生成摘要")
        
        # 3. 过滤低价值流程性记录（只在较短且命中噪音短语时过滤）
        NOISE_PHRASES = ['确认转写', '确认文本', '测试录音', '测试一下', '这是一条测试', '请点击链接']
        MIN_RECORD_LENGTH = 20  # 短记录阈值
        
        filtered_records = []
        for r in records:
            content = r.content or ""
            is_short = len(content) < MIN_RECORD_LENGTH
            has_noise = any(phrase in content for phrase in NOISE_PHRASES)
            # 只有较短且包含噪音短语的才过滤
            if is_short and has_noise:
                continue
            filtered_records.append(r)
        
        # 如果过滤后太少，保留原样
        if len(filtered_records) >= 2:
            records = filtered_records
            logger.info(f"[Summary] Filtered records: {len(records)} remain")
        
        # 4. 准备 records 文本（带序号，便于 LLM 理解演进）
        records_text = "\n---\n".join([
            f"【记录 {i+1} - {r.created_at.strftime('%Y-%m-%d')}】\n{r.content}"
            for i, r in enumerate(records)
        ])
        
        # 5. 更新状态为 updating
        customer.summary_status = "updating"
        await self.session.flush()
        logger.info(f"[Summary] Customer {customer_id} status -> updating, records: {len(records)}")
        
        # 6. 调用 LLM 生成摘要
        industry_key = await self._get_user_industry_key(user_id)
        prompt = _customer_summary_prompt(records_text, industry_key=industry_key)

        try:
            kimi = KimiClient()
            try:
                response = await kimi.chat_simple(
                    prompt=prompt,
                    system_prompt=_customer_summary_system(industry_key=industry_key),
                )
            finally:
                await kimi.close()

            if not response or not response.strip():
                raise ValueError("LLM 返回空摘要")

            summary_text = response.strip()
            
        except Exception as e:
            # 失败处理：更新状态为 failed，保留旧摘要
            logger.error(f"[Summary] Generation failed for customer {customer_id}: {e}")
            customer.summary_status = "failed"
            await self.session.flush()
            
            raise HTTPException(
                status_code=500,
                detail=f"摘要生成失败: {str(e)}"
            )
        
        # 7. 成功：保存新摘要，更新状态为 ready
        now = datetime.now()
        customer.summary_text = summary_text
        customer.summary_status = "ready"
        customer.updated_at = now
        await self.session.flush()
        
        logger.info(f"[Summary] Customer {customer_id} status -> ready")
        
        return SummaryGenerateResponse(
            customer_id=customer_id,
            summary_text=summary_text,
            summary_status="ready",
            records_count=len(records),
            updated_at=now,
        )
    
    async def chat_with_customer(
        self,
        user_id: str,
        customer_id: str,
        question: str,
    ) -> CustomerChatResponse:
        """
        基于客户摘要和最近记录回答问题
        
        Args:
            user_id: 当前用户ID
            customer_id: 客户ID
            question: 用户问题
            
        Returns:
            对话响应
            
        Raises:
            HTTPException: 校验失败或摘要不可用时抛出
        """
        from fastapi import HTTPException
        
        # 1. 校验客户存在且属于当前用户
        query = select(Customer).where(
            and_(
                Customer.id == customer_id,
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None)
            )
        )
        result = await self.session.execute(query)
        customer = result.scalar_one_or_none()
        
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在或无权访问")
        
        # 2. 查询最近 3 条 records（按时间倒序）
        records_query = select(Record).where(
            Record.customer_id == customer_id
        ).order_by(
            desc(Record.created_at)
        ).limit(3)
        
        result = await self.session.execute(records_query)
        records = result.scalars().all()
        
        # 4. 组装 records 文本（按时间正序，从旧到新）
        records_text = "\n\n".join([
            f"[{r.created_at.strftime('%Y-%m-%d')}] {r.content}"
            for r in reversed(records)
        ])

        summary_ready = bool(customer.summary_text and customer.summary_status == "ready")
        if not summary_ready and not records_text:
            return CustomerChatResponse(
                customer_id=customer_id,
                question=question,
                answer="当前记录中没有足够信息回答此问题。请先添加拜访记录，或生成客户画像后再问。",
            )

        summary_text = customer.summary_text if summary_ready else (
            "客户画像尚未生成。以下回答只能基于最近沟通记录，不代表完整客户画像。"
        )
        
        # 5. 组装 Prompt
        industry_key = await self._get_user_industry_key(user_id)
        prompt = _customer_chat_prompt(
            customer_summary_text=summary_text,
            recent_records_text=records_text,
            question=question,
            industry_key=industry_key,
        )

        # 6. 调用 Kimi 生成回答
        try:
            kimi = KimiClient()
            try:
                answer = await kimi.chat_simple(
                    prompt=prompt,
                    system_prompt=_customer_chat_system(industry_key=industry_key),
                )
            finally:
                await kimi.close()
            
            # 校验返回内容
            if not answer or not answer.strip():
                raise HTTPException(status_code=500, detail="AI 返回空回答")
                
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"[Chat] Kimi API failed for customer {customer_id}: {e}")
            raise HTTPException(status_code=500, detail=f"对话生成失败: {str(e)}")
        
        logger.info(f"[Chat] Customer {customer_id} question answered, length: {len(answer)}")
        
        return CustomerChatResponse(
            customer_id=customer_id,
            question=question,
            answer=answer.strip(),
        )
    
    async def generate_advice(
        self,
        user_id: str,
        customer_id: str,
    ) -> AdviceGenerateResponse:
        """
        基于客户摘要和最近记录生成跟进建议
        
        Args:
            user_id: 当前用户ID
            customer_id: 客户ID
            
        Returns:
            跟进建议响应
            
        Raises:
            HTTPException: 校验失败或摘要不可用时抛出
        """
        from fastapi import HTTPException
        
        # 1. 校验客户存在且属于当前用户
        query = select(Customer).where(
            and_(
                Customer.id == customer_id,
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None)
            )
        )
        result = await self.session.execute(query)
        customer = result.scalar_one_or_none()
        
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在或无权访问")
        
        # 2. 校验摘要是否可用（必须已生成且状态为 ready）
        if not customer.summary_text or customer.summary_status != "ready":
            raise HTTPException(
                status_code=400, 
                detail="客户摘要尚未生成，请先生成摘要"
            )
        
        # 3. 查询最近 5 条 records（按时间倒序）
        records_query = select(Record).where(
            Record.customer_id == customer_id
        ).order_by(
            desc(Record.created_at)
        ).limit(5)
        
        result = await self.session.execute(records_query)
        records = result.scalars().all()
        
        # 4. 组装 records 文本（按时间正序，从旧到新）
        records_text = "\n\n".join([
            f"[{r.created_at.strftime('%Y-%m-%d')}] {r.content}"
            for r in reversed(records)
        ])
        
        # 5. 组装 Prompt
        industry_key = await self._get_user_industry_key(user_id)
        prompt = _advice_prompt(
            customer_summary_text=customer.summary_text,
            recent_records_text=records_text,
            industry_key=industry_key,
        )

        # 6. 调用 Kimi 生成建议
        try:
            kimi = KimiClient()
            try:
                advice = await kimi.chat_simple(
                    prompt=prompt,
                    system_prompt=_advice_system(industry_key=industry_key),
                )
            finally:
                await kimi.close()
            
            # 校验返回内容
            if not advice or not advice.strip():
                raise HTTPException(status_code=500, detail="AI 返回空建议")
                
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"[Advice] Kimi API failed for customer {customer_id}: {e}")
            raise HTTPException(status_code=500, detail=f"建议生成失败: {str(e)}")
        
        logger.info(f"[Advice] Customer {customer_id} advice generated, length: {len(advice)}")
        
        saved = await self.advice_store.save_advice(
            user_id=user_id,
            customer_id=customer_id,
            advice_text=advice.strip(),
        )

        return AdviceGenerateResponse(
            customer_id=customer_id,
            advice_text=advice.strip(),
            updated_at=datetime.fromisoformat(saved["updated_at"]),
        )

    async def get_saved_advice(
        self,
        user_id: str,
        customer_id: str,
    ) -> Optional[AdviceGenerateResponse]:
        """读取已保存的拜访建议"""
        from fastapi import HTTPException

        query = select(Customer).where(
            and_(
                Customer.id == customer_id,
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None)
            )
        )
        result = await self.session.execute(query)
        customer = result.scalar_one_or_none()

        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在或无权访问")

        saved = await self.advice_store.get_advice(user_id=user_id, customer_id=customer_id)
        if not saved:
            return None

        updated_at = None
        if saved.get("updated_at"):
            updated_at = datetime.fromisoformat(saved["updated_at"])

        return AdviceGenerateResponse(
            customer_id=customer_id,
            advice_text=saved.get("advice_text", ""),
            updated_at=updated_at,
        )

    async def get_summary_stats(self, user_id: str) -> dict:
        """首页摘要统计：待更新画像数 + 超期未联系数 + 客户总数。"""
        from datetime import date as date_type
        from app.services.ai_service import STALE_CONTACT_MONTHS, subtract_months

        threshold = subtract_months(datetime.now().date(), STALE_CONTACT_MONTHS)
        threshold_dt = datetime(threshold.year, threshold.month, threshold.day)

        # 总数 + 按状态统计
        total_result = await self.session.execute(
            select(
                func.count(Customer.id),
                func.sum(
                    case(
                        (Customer.summary_status.in_(["stale", "failed"]), 1),
                        else_=0,
                    )
                ),
            ).where(
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None),
            )
        )
        total, stale_summary = total_result.one()
        total = total or 0
        stale_summary = stale_summary or 0

        # 超期未联系：最近一条 record 时间早于阈值 或 无任何记录
        latest_sub = (
            select(
                Record.customer_id,
                func.max(Record.created_at).label("last_contact"),
            )
            .group_by(Record.customer_id)
            .subquery()
        )

        stale_contact_result = await self.session.execute(
            select(func.count(Customer.id))
            .outerjoin(latest_sub, latest_sub.c.customer_id == Customer.id)
            .where(
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None),
                or_(
                    latest_sub.c.last_contact.is_(None),
                    latest_sub.c.last_contact < threshold_dt,
                ),
            )
        )
        stale_contact = stale_contact_result.scalar() or 0

        return {
            "customer_total": total,
            "stale_summary_count": stale_summary,
            "stale_contact_count": stale_contact,
        }

    async def export_customers_excel(self, user_id: str) -> bytes:
        """导出当前用户的客户资料为 Excel。"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        result = await self.session.execute(
            select(Customer)
            .where(
                Customer.user_id == user_id,
                Customer.deleted_at.is_(None),
            )
            .order_by(Customer.name.asc(), Customer.updated_at.desc())
        )
        customers = result.scalars().all()

        headers = [
            "客户姓名",
            "电话",
            "性别",
            "年龄",
            "生日",
            "地址",
            "标签",
            "客户画像",
            "下一步建议",
            "画像状态",
            "创建时间",
            "更新时间",
        ]
        status_labels = {
            "ready": "已生成",
            "stale": "待更新",
            "updating": "生成中",
            "failed": "生成失败",
        }

        wb = Workbook()
        ws = wb.active
        ws.title = "客户"
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="E7F5F2")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for customer in customers:
            address_parts = [
                customer.location_raw,
                customer.location_city,
                customer.location_district,
                customer.location_subarea,
            ]
            address = " / ".join(
                part.strip()
                for part in address_parts
                if isinstance(part, str) and part.strip()
            )
            ws.append(
                [
                    customer.name or "",
                    customer.phone or "",
                    customer.gender or "",
                    customer.age if customer.age is not None else "",
                    customer.birthday.strftime("%Y-%m-%d") if customer.birthday else "",
                    address,
                    "、".join(customer.tags or []),
                    customer.summary_text or "",
                    customer.advice_text or "",
                    status_labels.get(customer.summary_status, customer.summary_status or ""),
                    customer.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    if customer.created_at
                    else "",
                    customer.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                    if customer.updated_at
                    else "",
                ]
            )

        widths = [18, 18, 10, 8, 14, 28, 24, 48, 48, 12, 20, 20]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
